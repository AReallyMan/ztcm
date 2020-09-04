# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html
import os, sys
import zipfile

import dateparser
import time
from openpyxl import Workbook
from scrapy.mail import MailSender
from pymongo import MongoClient
from .settings import *


class ZzsendemailPipeline:
    def process_item(self, item, spider):
        return item


ccgp_items = [
    'title',
    'publish',
    'budget',
    'project',
    'type',
    'company',
    'noticeTime',
    'priceTime',
    'openTime',
    'address',
    'url',
    'docPrice',
    'docPlace',
    'project_contact',
    'project_telephone',
    'purchase_address',
    'purchase_telephone',
    'agency',
    'agency_address',
    'agency_telephone',
]


class ExcelPipeline(object):
    """
    1. 连接数据库
    2. 读取数据库集合
    3. 遍历集合每个集合生成一个excel文件 (指定路径下)
    4. 将所有excel文件打包 生成 zip    (带路径的名称)
    """

    def open_spider(self, spider):
        # 创建MONGODB数据库链接
        client = MongoClient(host=MONGODB_HOST, port=MONGODB_PORT)
        # 指定数据库
        self.mydb = client[MONGODB_DB]
        # 根据数据库集合生成的excel文件

        if 'linux' in sys.platform:
            self.exportpath = '/app/mongo_exportfile/' + spider.start_time + '---' + spider.end_time
        if 'win' in sys.platform:
            self.exportpath = os.path.join(os.path.abspath(os.path.join(os.getcwd(), "../..")),
                                           'mongo_exportfile/' + spider.start_time + '---' + spider.end_time)

        if not os.path.exists(self.exportpath):
            os.makedirs(self.exportpath)
        spider.zip_path = self.exportpath
        for file in os.listdir(self.exportpath):
            os.remove(os.path.join(self.exportpath, file))

    def process_item(self, item, spider):
        for coll in self.mydb.list_collection_names():
            # if in coll: 判断 ccgp ,caigoushunyi 格式和其他都不同
            colldata = self.mydb[coll]

            if spider.start_time == spider.end_time:
                filename = colldata.find()[0]['zh_name'] + '(' + spider.start_time + ').xlsx'
            else:
                filename = colldata.find()[0]['zh_name'] + '(' + spider.start_time + "--" + spider.end_time + ').xlsx'
            exportfile = os.path.join(self.exportpath, filename)
            if 'caigoushunyiSpider' in coll:
                try:
                    wb = Workbook()
                    ws = wb.active
                    ws.append(['标题', '预算金额', '采购项目名称', '品目',
                               '采购单位', '公告时间',
                               '获取招标文件时间', '招标文件发售地点', '投标截止时间', '开标时间', '开标地点', '链接'])
                    for row in colldata.find():
                        for i in ccgp_items:
                            try:
                                if not row[i]:
                                    row[i] = ''
                            except KeyError:
                                row[i] = ''
                        if self.judge_date_range(spider.start_time, spider.end_time, row['day']):
                            line = [row['title'], row['budget'], row['project'],
                                    row['type'], row['company'],
                                    row['noticeTime'], row['priceTime'], row['bidDocument'], row['bidDeadline'],
                                    row['openTime'], row['address'], row['url']]
                            ws.append(line)
                    wb.save(exportfile)
                    print(exportfile, '已经成功生成', 'caigoushunyiSpider')
                except Exception as e:
                    print(e)
                finally:
                    wb.close()

            elif 'ccgp' in coll:
                try:
                    wb = Workbook()
                    ws = wb.active
                    ws.append(['标题', '发布时间', '预算金额', '采购项目名称', '品目',
                               '采购单位', '招标文件售价', '获取招标文件地点', '公告时间',
                               '获取招标文件时间', '开标时间', '开标地点', '链接',
                               '项目联系人', '项目联系电话', '采购单位地址', '采购单位联系方式',
                               '代理机构名称', '代理机构地址', '代理机构联系方式'])
                    for row in colldata.find():
                        for i in ccgp_items:
                            try:
                                if not row[i]:
                                    row[i] = ''
                            except KeyError:
                                row[i] = ''
                        if self.judge_date_range(spider.start_time, spider.end_time, row['day']):
                            line = [row['title'], row['publish'], row['budget'], row['project'],
                                    row['type'], row['company'], row['docPrice'], row['docPlace'],
                                    row['noticeTime'], row['priceTime'],
                                    row['openTime'], row['address'], row['url'],
                                    row['project_contact'], row['project_telephone'], row['purchase_address'],
                                    row['purchase_telephone'], row['agency'], row['agency_address'],
                                    row['agency_telephone']]
                            ws.append(line)
                    wb.save(exportfile)
                    print(exportfile, '已经成功生成', 'ccgp')
                except Exception as e:
                    print(e)
                finally:
                    wb.close()

            else:
                try:
                    wb = Workbook()
                    ws = wb.active
                    ws.append(['分类', '标题', '发布时间', '内容', '链接'])
                    for row in colldata.find():
                        if self.judge_date_range(spider.start_time, spider.end_time, row['day']):
                            if row['content']:
                                content = row['content']
                            else:
                                content = ''
                            line = [row['type'], row['title'], row['day'], content, row['url']]
                            ws.append(line)
                    wb.save(exportfile)
                    print(exportfile, '已经成功生成', '其他')
                except Exception as e:
                    print(e)
                finally:
                    wb.close()

        # 压缩包路径
        spider.zip_path = self.createZipFile(spider.start_time, spider.end_time)
        return item

    # def format_excel(self, table_header, line)

    def judge_date_range(self, start_time, end_time, item_date):
        # s_time = int(dateparser.parse(start_time).timestamp())     # 00:00:00
        # # s_time = int(time.mktime(time.strptime(dt, "%Y-%m-%d %H:%M:%S")))     # 00:00:00
        # e_time = int(dateparser.parse(end_time).timestamp()) + 86399  # 23:59:59
        # item_date = dateparser.parse(item_date).timestamp()
        s_time = int(time.mktime(time.strptime(start_time, "%Y-%m-%d")))  # 00:00:00
        e_time = int(time.mktime(time.strptime(start_time, "%Y-%m-%d"))) + 86399  # 23:59:59
        item_date = int(time.mktime(time.strptime(item_date, "%Y-%m-%d")))
        if s_time <= item_date <= e_time:
            return True

    def createZipFile(self, s_time, e_time):
        if s_time == e_time:
            filename = os.path.join(self.exportpath, 'contract file(' + e_time + ').zip')
        else:
            filename = os.path.join(self.exportpath, 'contract file(' + s_time + "--" + e_time + ').zip')
        # os.chdir(self.exportpath)
        with zipfile.ZipFile(filename, 'w', zipfile.ZIP_STORED) as f:
            files = [os.path.join(self.exportpath, file) for file in os.listdir(self.exportpath) if
                     '.xlsx' in file]
            [f.write(file) for file in files]
        return filename


# 找到 zip 附件， 发送邮件
class EmailPipeline(object):

    def process_item(self, item, spider):
        return item

    def close_spider(self, spider):
        mailer = MailSender(smtphost=STMPHOST,
                            mailfrom=EMAIL_FROM,
                            smtppass=STMPPASS,
                            smtpuser=EMAIL_FROM,
                            smtpport=STMPPORT,
                            smtptls=True
                            )
        if spider.start_time == spider.end_time:
            subject = '(' + spider.end_time + ')招标文件，及时查收'
        else:
            subject = '(' + spider.start_time + '--' + spider.end_time + ')招标文件，及时查收'
        file = spider.zip_path
        if os.path.isfile(file):
            print(type(os.path.basename(file)))
            attachs = [(os.path.basename(file), EMAIL_ATTACH_MIME, open(file, "rb"))]
            body = '招标邮件，及时查收'.encode('utf-8')
        else:
            body = '今日无数据'.encode('utf-8')
            attachs = ()
        return mailer.send(to=EMAIL_TO, subject=subject, body=body,
                           cc=["zhzwx9@163.com"],
                           attachs=attachs, mimetype="text/plain", charset='utf-8')
